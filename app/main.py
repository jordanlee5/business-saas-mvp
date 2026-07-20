from fastapi import FastAPI, Request, Form, UploadFile, File, Query
from fastapi.responses import (
    HTMLResponse,
    RedirectResponse,
    FileResponse,
    StreamingResponse,
)
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import or_

import os
import hashlib
import io
import zipfile
from datetime import datetime, date, time
from decimal import Decimal, ROUND_HALF_UP
from urllib.parse import urlencode
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from .database import engine, Base, SessionLocal
from . import models
from .models import User, BusinessRecord, UploadBatch, VoucherRecord, VoucherUploadBatch ,MatchReview
from .auth import verify_password, get_password_hash
from .excel_service import parse_business_excel
from .ocr_service import ocr_image, match_ocr_with_records, extract_voucher_amount
from .business_no import generate_public_business_no
from .settlement_calculator import (
    EXTERNAL_MODE,
    INTERNAL_MODE,
    calculate_business_settlement,
)

app = FastAPI(title="业务数据管理SaaS MVP")

# 创建数据库表
Base.metadata.create_all(bind=engine)

os.makedirs("app/static", exist_ok=True)
app.mount("/static", StaticFiles(directory="app/static"), name="static")

os.makedirs("uploads", exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

templates = Jinja2Templates(directory="app/templates")

ACCEPTED_BATCH_STATUS = "已承接"


def generate_unique_public_business_no(
    db,
    reserved_numbers=None,
    max_attempts=100,
):
    """
    生成一个数据库中尚未使用的公开业务单号。

    reserved_numbers 用于防止同一次批量上传过程中，
    尚未提交的多条业务意外获得相同编号。
    """
    if reserved_numbers is None:
        reserved_numbers = set()

    for _ in range(max_attempts):
        candidate = generate_public_business_no()

        if candidate in reserved_numbers:
            continue

        exists = (
            db.query(BusinessRecord.id)
            .filter(
                BusinessRecord.public_business_no == candidate
            )
            .first()
        )

        if not exists:
            reserved_numbers.add(candidate)
            return candidate

    raise RuntimeError(
        "多次生成公开业务单号均发生重复，请停止操作并检查。"
    )


def apply_accepted_batch_filter(query):
    return (
        query
        .join(UploadBatch, BusinessRecord.batch_id == UploadBatch.id)
        .filter(UploadBatch.acceptance_status == ACCEPTED_BATCH_STATUS)
    )

def clean_zip_filename_part(value, fallback):
    """
    清理 ZIP 内部文件名组成部分。

    不修改数据库和磁盘上的原文件名，
    仅用于生成用户下载后看到的规范名称。
    """
    text = str(value or "").strip()

    invalid_characters = '<>:"/\\|?*'

    for invalid_character in invalid_characters:
        text = text.replace(invalid_character, "_")

    text = (
        text
        .replace("\r", "_")
        .replace("\n", "_")
        .replace("\t", "_")
    )

    while "__" in text:
        text = text.replace("__", "_")

    # Windows 文件名不适合以空格或英文句点结尾。
    text = text.strip(" ._")

    return text or fallback


def build_voucher_zip_filename(
    sequence,
    business_record,
    voucher,
    review,
    absolute_file_path,
):
    """
    统一生成 ZIP 内的凭证文件名。

    命名规范：
    序号_客户姓名_公开业务单号_凭证金额_审核记录ID.扩展名
    """
    customer_name = clean_zip_filename_part(
        business_record.name,
        "未知姓名",
    )

    business_no_part = clean_zip_filename_part(
        business_record.display_business_no,
        f"business_{business_record.id}",
    )

    voucher_amount_text = (
        f"{float(voucher.voucher_amount or 0):.2f}"
    )

    review_id_text = f"MR{review.id}"

    file_extension = os.path.splitext(
        absolute_file_path
    )[1].lower()

    if not file_extension:
        file_extension = ".bin"

    return (
        f"{sequence:03d}_"
        f"{customer_name}_"
        f"{business_no_part}_"
        f"{voucher_amount_text}_"
        f"{review_id_text}"
        f"{file_extension}"
    )

def money2(value):
    if value is None:
        return 0.0

    try:
        return float(
            Decimal(str(value)).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )
        )
    except Exception:
        return 0.0


def get_current_user(request: Request):
    user_id = request.cookies.get("user_id")
    if not user_id:
        return None

    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == int(user_id)).first()
        return user
    except Exception:
        return None
    finally:
        db.close()


def add_base_context(request: Request, context: dict):
    user = get_current_user(request)

    if user:
        context["username"] = user.username
        context["role"] = user.role
        context["topbar_username"] = user.username
        context["topbar_role"] = user.role
    else:
        context["username"] = ""
        context["role"] = ""
        context["topbar_username"] = ""
        context["topbar_role"] = ""

    return context


@app.get("/vouchers/{review_id}/download")
def download_approved_voucher(
    request: Request,
    review_id: int,
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    db = SessionLocal()

    try:
        review = db.query(MatchReview).filter(MatchReview.id == review_id).first()

        if not review:
            return RedirectResponse(url="/business-records", status_code=302)

        # 下载入口只允许下载“已通过”的凭证
        if review.review_status != "已通过":
            return RedirectResponse(url="/business-records", status_code=302)

        business_record = (
            db.query(BusinessRecord)
            .filter(BusinessRecord.id == review.business_record_id)
            .first()
        )

        voucher = (
            db.query(VoucherRecord)
            .filter(VoucherRecord.id == review.voucher_id)
            .first()
        )

        if not business_record or not voucher:
            return RedirectResponse(url="/business-records", status_code=302)

        # 权限隔离：
        # 管理员可以下载已通过凭证；
        # 上传方只能下载自己业务数据关联的已通过凭证。
        if user.role == "partner" and business_record.user_id != user.id:
            return RedirectResponse(url="/dashboard", status_code=302)

        if user.role not in ["admin", "partner"]:
            return RedirectResponse(url="/dashboard", status_code=302)

        if not voucher.file_path:
            return RedirectResponse(url="/business-records", status_code=302)

        relative_path = voucher.file_path.replace("\\", "/").lstrip("/")
        absolute_file_path = os.path.abspath(relative_path)
        uploads_root = os.path.abspath("uploads")

        # 防止下载 uploads 目录外的文件
        if os.path.commonpath([uploads_root, absolute_file_path]) != uploads_root:
            return RedirectResponse(url="/business-records", status_code=302)

        if not os.path.exists(absolute_file_path):
            return RedirectResponse(url="/business-records", status_code=302)

        download_filename = voucher.filename or os.path.basename(absolute_file_path)

        return FileResponse(
            path=absolute_file_path,
            filename=download_filename,
            media_type="application/octet-stream",
        )

    finally:
        db.close()


@app.get("/business-records/{record_id}/vouchers/download-all")
def download_all_approved_vouchers(
    request: Request,
    record_id: int,
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # 当前批量下载功能先只开放给上传方。
    if user.role != "partner":
        return RedirectResponse(
            url=f"/business-records/{record_id}",
            status_code=302,
        )

    db = SessionLocal()

    try:
        business_record = (
            db.query(BusinessRecord)
            .filter(BusinessRecord.id == record_id)
            .first()
        )

        if not business_record:
            return RedirectResponse(
                url="/business-records",
                status_code=302,
            )

        # 上传方只能批量下载自己名下业务的凭证。
        if business_record.user_id != user.id:
            return RedirectResponse(
                url="/dashboard",
                status_code=302,
            )

        approved_reviews = (
            db.query(MatchReview)
            .filter(MatchReview.business_record_id == business_record.id)
            .filter(MatchReview.review_status == "已通过")
            .order_by(MatchReview.id.asc())
            .all()
        )

        if not approved_reviews:
            return RedirectResponse(
                url=f"/business-records/{record_id}",
                status_code=302,
            )

        uploads_root = os.path.abspath("uploads")
        zip_buffer = io.BytesIO()

        added_file_count = 0
        seen_voucher_ids = set()


        with zipfile.ZipFile(
            zip_buffer,
            mode="w",
            compression=zipfile.ZIP_DEFLATED,
        ) as zip_file:

            for review in approved_reviews:
                # 防止同一凭证因重复审核记录被加入两次。
                if review.voucher_id in seen_voucher_ids:
                    continue

                seen_voucher_ids.add(review.voucher_id)

                voucher = (
                    db.query(VoucherRecord)
                    .filter(VoucherRecord.id == review.voucher_id)
                    .first()
                )

                if not voucher or not voucher.file_path:
                    continue

                relative_path = (
                    voucher.file_path
                    .replace("\\", "/")
                    .lstrip("/")
                )

                absolute_file_path = os.path.abspath(relative_path)

                # 防止访问 uploads 目录之外的文件。
                try:
                    common_path = os.path.commonpath(
                        [uploads_root, absolute_file_path]
                    )
                except ValueError:
                    continue

                if common_path != uploads_root:
                    continue

                if not os.path.isfile(absolute_file_path):
                    continue

                added_file_count += 1

                archive_filename = build_voucher_zip_filename(
                    sequence=added_file_count,
                    business_record=business_record,
                    voucher=voucher,
                    review=review,
                    absolute_file_path=absolute_file_path,
                )

                zip_file.write(
                    absolute_file_path,
                    arcname=archive_filename,
                )

        # 可能存在数据库记录，但对应文件全部丢失的情况。
        if added_file_count == 0:
            return RedirectResponse(
                url=f"/business-records/{record_id}",
                status_code=302,
            )

        zip_buffer.seek(0)

        business_no = (
            business_record.display_business_no
            or f"business_{business_record.id}"
        )

        # ZIP 外层文件名先使用业务单号，避免中文响应头兼容问题。
        safe_business_no = "".join(
            character
            for character in business_no
            if character.isalnum() or character in ("-", "_")
        )

        if not safe_business_no:
            safe_business_no = f"business_{business_record.id}"

        download_filename = (
            f"{safe_business_no}_approved_vouchers.zip"
        )

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": (
                    f'attachment; filename="{download_filename}"'
                )
            },
        )

    finally:
        db.close()


@app.get("/my-settlement/vouchers/download")
def download_my_settlement_vouchers(
    request: Request,
    start_date: str = Query(""),
    end_date: str = Query(""),
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(
            url="/login",
            status_code=302,
        )

    # 第一版只开放给上传方。
    if user.role != "partner":
        return RedirectResponse(
            url="/dashboard",
            status_code=302,
        )

    redirect_query = urlencode(
        {
            "start_date": start_date,
            "end_date": end_date,
        }
    )
    settlement_url = f"/my-settlement?{redirect_query}"

    start_datetime = None
    end_datetime = None

    try:
        if start_date:
            start_datetime = datetime.combine(
                datetime.strptime(
                    start_date,
                    "%Y-%m-%d",
                ).date(),
                time.min,
            )

        if end_date:
            end_datetime = datetime.combine(
                datetime.strptime(
                    end_date,
                    "%Y-%m-%d",
                ).date(),
                time.max,
            )

    except ValueError:
        return RedirectResponse(
            url=settlement_url,
            status_code=302,
        )

    if (
        start_datetime
        and end_datetime
        and start_datetime > end_datetime
    ):
        return RedirectResponse(
            url=settlement_url,
            status_code=302,
        )

    db = SessionLocal()

    try:
        # 与“我的结算报表”保持一致：
        # 只查询已承接批次下、当前 Partner 自己的业务。
        records_query = apply_accepted_batch_filter(
            db.query(BusinessRecord)
        ).filter(
            BusinessRecord.user_id == user.id
        )

        if start_datetime:
            records_query = records_query.filter(
                BusinessRecord.created_at >= start_datetime
            )

        if end_datetime:
            records_query = records_query.filter(
                BusinessRecord.created_at <= end_datetime
            )

        business_records = (
            records_query
            .order_by(BusinessRecord.id.asc())
            .all()
        )

        if not business_records:
            return RedirectResponse(
                url=settlement_url,
                status_code=302,
            )

        uploads_root = os.path.abspath("uploads")
        zip_buffer = io.BytesIO()
        total_added_file_count = 0

        with zipfile.ZipFile(
            zip_buffer,
            mode="w",
            compression=zipfile.ZIP_DEFLATED,
        ) as zip_file:

            for business_record in business_records:
                # 与结算页当前口径一致：
                # 最新审核结果为“已通过”才属于已通过结算业务。
                latest_review = (
                    db.query(MatchReview)
                    .filter(
                        MatchReview.business_record_id
                        == business_record.id
                    )
                    .order_by(MatchReview.id.desc())
                    .first()
                )

                if (
                    not latest_review
                    or latest_review.review_status != "已通过"
                ):
                    continue

                approved_reviews = (
                    db.query(MatchReview)
                    .filter(
                        MatchReview.business_record_id
                        == business_record.id
                    )
                    .filter(
                        MatchReview.review_status == "已通过"
                    )
                    .order_by(MatchReview.id.asc())
                    .all()
                )

                if not approved_reviews:
                    continue

                folder_name = clean_zip_filename_part(
                    business_record.display_business_no,
                    f"business_{business_record.id}",
                )

                # 每个业务文件夹独立去重并重新从 001 编号。
                seen_voucher_ids = set()
                business_file_count = 0

                for review in approved_reviews:
                    if review.voucher_id in seen_voucher_ids:
                        continue

                    seen_voucher_ids.add(review.voucher_id)

                    voucher = (
                        db.query(VoucherRecord)
                        .filter(
                            VoucherRecord.id
                            == review.voucher_id
                        )
                        .first()
                    )

                    if not voucher or not voucher.file_path:
                        continue

                    relative_path = (
                        voucher.file_path
                        .replace("\\", "/")
                        .lstrip("/")
                    )

                    absolute_file_path = os.path.abspath(
                        relative_path
                    )

                    # 防止 ZIP 下载访问 uploads 目录之外的文件。
                    try:
                        common_path = os.path.commonpath(
                            [
                                uploads_root,
                                absolute_file_path,
                            ]
                        )
                    except ValueError:
                        continue

                    if common_path != uploads_root:
                        continue

                    if not os.path.isfile(absolute_file_path):
                        continue

                    business_file_count += 1
                    total_added_file_count += 1

                    archive_filename = build_voucher_zip_filename(
                        sequence=business_file_count,
                        business_record=business_record,
                        voucher=voucher,
                        review=review,
                        absolute_file_path=absolute_file_path,
                    )

                    zip_file.write(
                        absolute_file_path,
                        arcname=(
                            f"{folder_name}/"
                            f"{archive_filename}"
                        ),
                    )

        if total_added_file_count == 0:
            return RedirectResponse(
                url=settlement_url,
                status_code=302,
            )

        zip_buffer.seek(0)

        start_date_part = (
            start_date.replace("-", "")
            if start_date
            else "all"
        )

        end_date_part = (
            end_date.replace("-", "")
            if end_date
            else "all"
        )

        download_filename = (
            "my_settlement_vouchers_"
            f"{start_date_part}_"
            f"{end_date_part}.zip"
        )

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": (
                    f'attachment; filename="{download_filename}"'
                )
            },
        )

    finally:
        db.close()


def get_voucher_batch_review_summary(db, batch_id: int):
    # 兼容不同字段名：如果你的 VoucherRecord 里叫 voucher_batch_id / batch_id / upload_batch_id，都能识别
    batch_column = None

    if hasattr(VoucherRecord, "voucher_batch_id"):
        batch_column = VoucherRecord.voucher_batch_id
    elif hasattr(VoucherRecord, "batch_id"):
        batch_column = VoucherRecord.batch_id
    elif hasattr(VoucherRecord, "upload_batch_id"):
        batch_column = VoucherRecord.upload_batch_id

    if batch_column is None:
        return {
            "total_review_count": 0,
            "pending_review_count": 0,
            "finished_review_count": 0,
            "action_type": "none",
            "action_text": "无匹配记录",
        }

    vouchers = db.query(VoucherRecord).filter(batch_column == batch_id).all()
    voucher_ids = [voucher.id for voucher in vouchers]

    if not voucher_ids:
        return {
            "total_review_count": 0,
            "pending_review_count": 0,
            "finished_review_count": 0,
            "action_type": "none",
            "action_text": "无匹配记录",
        }

    total_review_count = (
        db.query(MatchReview)
        .filter(MatchReview.voucher_id.in_(voucher_ids))
        .count()
    )

    pending_review_count = (
        db.query(MatchReview)
        .filter(MatchReview.voucher_id.in_(voucher_ids))
        .filter(MatchReview.review_status == "待审核")
        .count()
    )

    finished_review_count = total_review_count - pending_review_count

    if total_review_count == 0:
        action_type = "none"
        action_text = "无匹配记录"
    elif pending_review_count > 0:
        action_type = "pending"
        action_text = "去审核"
    else:
        action_type = "finished"
        action_text = "查看结果"

    return {
        "total_review_count": total_review_count,
        "pending_review_count": pending_review_count,
        "finished_review_count": finished_review_count,
        "action_type": action_type,
        "action_text": action_text,
    }


def build_stats_data(partner_id: int = 0, start_date: str = "", end_date: str = ""):
    db = SessionLocal()

    partners = db.query(User).filter(User.role == "partner").order_by(User.id.desc()).all()

    selected_partner_id = partner_id
    selected_start_date = start_date
    selected_end_date = end_date

    start_datetime = None
    end_datetime = None

    try:
        if selected_start_date:
            start_datetime = datetime.combine(
                datetime.strptime(selected_start_date, "%Y-%m-%d").date(),
                time.min,
            )

        if selected_end_date:
            end_datetime = datetime.combine(
                datetime.strptime(selected_end_date, "%Y-%m-%d").date(),
                time.max,
            )
    except Exception:
        start_datetime = None
        end_datetime = None
        selected_start_date = ""
        selected_end_date = ""

    # 统计/核算口径：只统计已承接批次下的业务数据
    records_query = apply_accepted_batch_filter(db.query(BusinessRecord))
    batches_query = db.query(UploadBatch).filter(
        UploadBatch.acceptance_status == ACCEPTED_BATCH_STATUS
    )

    if start_datetime:
        records_query = records_query.filter(BusinessRecord.created_at >= start_datetime)
        batches_query = batches_query.filter(UploadBatch.created_at >= start_datetime)

    if end_datetime:
        records_query = records_query.filter(BusinessRecord.created_at <= end_datetime)
        batches_query = batches_query.filter(UploadBatch.created_at <= end_datetime)

    if selected_partner_id != 0:
        records_query = records_query.filter(BusinessRecord.user_id == selected_partner_id)
        batches_query = batches_query.filter(UploadBatch.user_id == selected_partner_id)

    business_records = records_query.all()
    accepted_record_ids = [record.id for record in business_records]

    total_records = len(business_records)
    total_batches = batches_query.count()
    total_partners = db.query(User).filter(User.role == "partner").count()

    reviews_query = db.query(MatchReview)

    # 匹配审核统计也只统计已承接批次下的业务数据
    if accepted_record_ids:
        reviews_query = reviews_query.filter(
            MatchReview.business_record_id.in_(accepted_record_ids)
        )
    else:
        reviews_query = reviews_query.filter(MatchReview.id == -1)

    if start_datetime:
        reviews_query = reviews_query.filter(MatchReview.created_at >= start_datetime)

    if end_datetime:
        reviews_query = reviews_query.filter(MatchReview.created_at <= end_datetime)


    pending_reviews = reviews_query.filter(MatchReview.review_status == "待审核").count()
    approved_reviews = reviews_query.filter(MatchReview.review_status == "已通过").count()
    rejected_reviews = reviews_query.filter(MatchReview.review_status == "已驳回").count()

    total_points = 0

    # 费用按单条业务先精确到分，再使用 Decimal 汇总。
    total_receivable_fee = Decimal("0.00")
    total_payable_cost = Decimal("0.00")
    total_gross_profit = Decimal("0.00")

    approved_settlement_count = 0
    approved_settlement_points = 0
    approved_settlement_receivable_fee = Decimal(
        "0.00"
    )
    approved_settlement_payable_cost = Decimal(
        "0.00"
    )
    approved_settlement_gross_profit = Decimal(
        "0.00"
    )

    rows = []

    for record in business_records:
        uploader = db.query(User).filter(User.id == record.user_id).first()

        points_amount = record.points_amount or 0
        service_rate = record.record_service_rate if record.record_service_rate is not None else 0
        upstream_cost_rate = (
            record.record_upstream_cost_rate
            if record.record_upstream_cost_rate is not None
            else 0
        )

        service_rate_mode = (
            record.record_service_rate_mode
            if record.record_service_rate_mode
            in (
                "external",
                "internal",
            )
            else "external"
        )

        upstream_cost_rate_mode = (
            record.record_upstream_cost_rate_mode
            if record.record_upstream_cost_rate_mode
            in (
                "external",
                "internal",
            )
            else "external"
        )

        settlement_result = calculate_business_settlement(
            base_amount=points_amount,
            downstream_rate_percent=service_rate,
            downstream_mode=service_rate_mode,
            upstream_rate_percent=upstream_cost_rate,
            upstream_mode=upstream_cost_rate_mode,
        )

        receivable_fee = (
            settlement_result.downstream.fee_amount
        )

        payable_cost = (
            settlement_result.upstream.fee_amount
        )

        gross_profit = settlement_result.gross_profit

        total_points += points_amount
        total_receivable_fee += receivable_fee
        total_payable_cost += payable_cost
        total_gross_profit += gross_profit

        latest_review = (
            db.query(MatchReview)
            .filter(MatchReview.business_record_id == record.id)
            .order_by(MatchReview.id.desc())
            .first()
        )

        if latest_review and latest_review.review_status == "已通过":
            approved_settlement_count += 1
            approved_settlement_points += points_amount
            approved_settlement_receivable_fee += receivable_fee
            approved_settlement_payable_cost += payable_cost
            approved_settlement_gross_profit += gross_profit

        review = (
            db.query(MatchReview)
            .filter(MatchReview.business_record_id == record.id)
            .order_by(MatchReview.id.desc())
            .first()
        )

        review_status = latest_review.review_status if latest_review else "未匹配"

        rows.append(
            {
                "上传方": uploader.username if uploader else "未知上传方",
                "业务单号": record.display_business_no,
                "姓名": record.name,
                "手机号": record.phone,
                "车牌号": record.plate_number,
                "积分金额": points_amount,
                "银行卡号": record.bank_card,
                "下游服务费率": service_rate,
                "应收服务费": float(receivable_fee),
                "上游成本费率": upstream_cost_rate,
                "应付成本费": float(payable_cost),
                "毛利": float(gross_profit),
                "审核状态": review_status,
                "导入时间": record.created_at,
            }
        )

    db.close()

    return {
        "partners": partners,
        "selected_partner_id": selected_partner_id,
        "selected_start_date": selected_start_date,
        "selected_end_date": selected_end_date,
        "total_records": total_records,
        "total_batches": total_batches,
        "total_partners": total_partners,
        "pending_reviews": pending_reviews,
        "approved_reviews": approved_reviews,
        "rejected_reviews": rejected_reviews,
        "total_points": round(total_points, 2),
        "total_receivable_fee": float(
            total_receivable_fee
        ),
        "total_payable_cost": float(
            total_payable_cost
        ),
        "total_gross_profit": float(
            total_gross_profit
        ),

        "approved_settlement_count": approved_settlement_count,
        "approved_settlement_points": round(approved_settlement_points, 2),
        "approved_settlement_receivable_fee": float(
            approved_settlement_receivable_fee
        ),
        "approved_settlement_payable_cost": float(
            approved_settlement_payable_cost
        ),
        "approved_settlement_gross_profit": float(
            approved_settlement_gross_profit
        ),

        "rows": rows,
    }

def format_excel_file(writer):
    workbook = writer.book

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                cell_value = cell.value
                if cell_value is None:
                    continue

                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length

                if isinstance(cell_value, (int, float)):
                    cell.number_format = "#,##0.00"

                if "时间" in str(worksheet.cell(row=1, column=cell.column).value):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"

            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 35)

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    user = get_current_user(request)

    if user:
        return RedirectResponse(url="/dashboard", status_code=302)

    return RedirectResponse(url="/login", status_code=302)


@app.get("/query-record", response_class=HTMLResponse)
def query_record_page(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse(
        "query_record.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "business_records",
            "records": None,
            "keyword": "",
        },
    )


@app.post("/query-record", response_class=HTMLResponse)
def query_record_submit(
    request: Request,
    keyword: str = Form(...),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    keyword = keyword.strip()

    db = SessionLocal()

    query = db.query(BusinessRecord)

    # 权限隔离：
    # 管理员可以查所有数据
    # 上传方只能查自己上传的数据
    if user.role != "admin":
        query = query.filter(BusinessRecord.user_id == user.id)

    query_result = query.filter(
        or_(
            BusinessRecord.business_no == keyword,
            BusinessRecord.public_business_no == keyword,
            BusinessRecord.name == keyword,
            BusinessRecord.phone == keyword,
            BusinessRecord.plate_number == keyword,
            BusinessRecord.bank_card == keyword,
        )
    ).order_by(BusinessRecord.id.desc()).all()

    records = []

    for r in query_result:
        uploader = db.query(User).filter(User.id == r.user_id).first()

        service_rate = r.record_service_rate if r.record_service_rate is not None else 0
        upstream_cost_rate = r.record_upstream_cost_rate if r.record_upstream_cost_rate is not None else 0

        points_amount = r.points_amount or 0

        service_rate_mode = (
            r.record_service_rate_mode
            if r.record_service_rate_mode
            in (
                "external",
                "internal",
            )
            else "external"
        )

        upstream_cost_rate_mode = (
            r.record_upstream_cost_rate_mode
            if r.record_upstream_cost_rate_mode
            in (
                "external",
                "internal",
            )
            else "external"
        )

        settlement_result = calculate_business_settlement(
            base_amount=points_amount,
            downstream_rate_percent=service_rate,
            downstream_mode=service_rate_mode,
            upstream_rate_percent=upstream_cost_rate,
            upstream_mode=upstream_cost_rate_mode,
        )

        receivable_fee = (
            settlement_result.downstream.fee_amount
        )

        payable_cost = (
            settlement_result.upstream.fee_amount
        )

        gross_profit = settlement_result.gross_profit

        records.append(
            {
                "id": r.id,
                "business_no": r.display_business_no,
                "name": r.name,
                "phone": r.phone,
                "plate_number": r.plate_number,
                "points_amount": points_amount,
                "bank_card": r.bank_card,
                "created_at": r.created_at,
                "uploader_username": uploader.username if uploader else "未知上传方",
                "service_rate": service_rate,
                "upstream_cost_rate": upstream_cost_rate,
                "receivable_fee": float(receivable_fee),
                "payable_cost": float(payable_cost),
                "gross_profit": float(gross_profit),
            }
        )

    db.close()

    return templates.TemplateResponse(
        "query_record.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "business_records",
            "records": records,
            "keyword": keyword,
        },
    )


def build_business_record_items(
    db,
    user,
    partner_id=0,
    keyword="",
    start_date="",
    end_date="",
    review_status="全部",
    acceptance_filter="全部",
    page=1,
    page_size=10,
    use_pagination=True,
):
    query = db.query(BusinessRecord)

    # 权限隔离：管理员看全部；上传方只能看自己的
    if user.role == "partner":
        query = query.filter(BusinessRecord.user_id == user.id)
    else:
        if partner_id != 0:
            query = query.filter(BusinessRecord.user_id == partner_id)

    keyword = keyword.strip()

    if keyword:
        query = query.filter(
            or_(
                BusinessRecord.business_no.contains(keyword),
                BusinessRecord.public_business_no.contains(keyword),
                BusinessRecord.name.contains(keyword),
                BusinessRecord.phone.contains(keyword),
                BusinessRecord.plate_number.contains(keyword),
                BusinessRecord.bank_card.contains(keyword),
            )
        )

    if start_date:
        start_dt = datetime.combine(datetime.strptime(start_date, "%Y-%m-%d").date(), time.min)
        query = query.filter(BusinessRecord.created_at >= start_dt)

    if end_date:
        end_dt = datetime.combine(datetime.strptime(end_date, "%Y-%m-%d").date(), time.max)
        query = query.filter(BusinessRecord.created_at <= end_dt)

    all_records = query.order_by(BusinessRecord.id.desc()).all()

    all_record_items = []

    for record in all_records:
        uploader = db.query(User).filter(User.id == record.user_id).first()

        batch = None
        if record.batch_id:
            batch = db.query(UploadBatch).filter(UploadBatch.id == record.batch_id).first()

        acceptance_status = batch.acceptance_status if batch and batch.acceptance_status else "待承接"

        latest_review = (
            db.query(MatchReview)
            .filter(MatchReview.business_record_id == record.id)
            .order_by(MatchReview.id.desc())
            .first()
        )

        if acceptance_status == "已拒绝":
            latest_review_status = "已拒绝承接"
            latest_match_status = "-"
        elif acceptance_status == "待承接":
            latest_review_status = "待承接"
            latest_match_status = "-"
        else:
            latest_review_status = "未审核"
            latest_match_status = "未匹配"

            if latest_review:
                latest_review_status = latest_review.review_status
                latest_match_status = latest_review.match_status

        approved_reviews = (
            db.query(MatchReview)
            .filter(MatchReview.business_record_id == record.id)
            .filter(MatchReview.review_status == "已通过")
            .all()
        )

        approved_voucher_amount = 0

        for approved_review in approved_reviews:
            voucher = (
                db.query(VoucherRecord)
                .filter(VoucherRecord.id == approved_review.voucher_id)
                .first()
            )

            if voucher and voucher.voucher_amount:
                approved_voucher_amount += voucher.voucher_amount

        business_amount = money2(record.points_amount)
        approved_voucher_amount = money2(approved_voucher_amount)
        remaining_amount = money2(business_amount - approved_voucher_amount)

        if remaining_amount < 0:
            remaining_amount = 0.0

        # 业务完成优先：
        # 如果已通过凭证金额已经覆盖业务金额，则业务列表的审核状态应展示为“已通过”，
        # 不再被后续生成的“待审核 / 无需审核”等过程记录覆盖。
        if (
            acceptance_status == "已承接"
            and business_amount > 0
            and approved_voucher_amount > 0
            and remaining_amount <= 0
        ):
            latest_review_status = "已通过"

        item = {
            "id": record.id,
            "business_no": record.display_business_no,
            "uploader_username": uploader.username if uploader else "未知上传方",
            "acceptance_status": acceptance_status,
            "name": record.name,
            "phone": record.phone,
            "plate_number": record.plate_number,
            "points_amount": business_amount,
            "bank_card": record.bank_card,
            "latest_review_status": latest_review_status,
            "latest_match_status": latest_match_status,
            "approved_voucher_amount": money2(approved_voucher_amount),
            "remaining_amount": money2(remaining_amount),
            "created_at": record.created_at,
        }

        if review_status != "全部" and item["latest_review_status"] != review_status:
            continue
        
        if acceptance_filter != "全部" and item["acceptance_status"] != acceptance_filter:
            continue

        all_record_items.append(item)

    total_records = len(all_record_items)

    if not use_pagination:
        return all_record_items, total_records, 1, 1

    if page < 1:
        page = 1

    allowed_page_sizes = [3, 5, 10, 20]

    if page_size not in allowed_page_sizes:
        page_size = 10

    total_pages = (total_records + page_size - 1) // page_size

    if total_pages == 0:
        total_pages = 1

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * page_size
    page_items = all_record_items[offset: offset + page_size]

    return page_items, total_records, total_pages, page


@app.get("/business-records", response_class=HTMLResponse)
def business_records_page(
    request: Request,
    partner_id: int = Query(0),
    keyword: str = Query(""),
    start_date: str = Query(""),
    end_date: str = Query(""),
    review_status: str = Query("全部"),
    acceptance_status: str = Query("全部"),
    page: int = Query(1),
    page_size: int = Query(3),
    batch_page: int = Query(1),
    batch_page_size: int = Query(3),
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    db = SessionLocal()

    allowed_acceptance_statuses = ["全部", "待承接", "已承接", "已拒绝"]

    if acceptance_status not in allowed_acceptance_statuses:
        acceptance_status = "全部"

    batch_query = db.query(UploadBatch)

    # 权限隔离：
    # 管理员可以看全部上传批；如果选择了上传方，则只看该上传方的上传批次
    # 上传方只能看自己的上传批次
    if user.role != "admin":
        batch_query = batch_query.filter(UploadBatch.user_id == user.id)
    else:
        if partner_id != 0:
            batch_query = batch_query.filter(UploadBatch.user_id == partner_id)

    if start_date:
        start_dt = datetime.combine(datetime.strptime(start_date, "%Y-%m-%d").date(), time.min)
        batch_query = batch_query.filter(UploadBatch.created_at >= start_dt)

    if end_date:
        end_dt = datetime.combine(datetime.strptime(end_date, "%Y-%m-%d").date(), time.max)
        batch_query = batch_query.filter(UploadBatch.created_at <= end_dt)

    if acceptance_status != "全部":
        batch_query = batch_query.filter(UploadBatch.acceptance_status == acceptance_status)

    allowed_batch_page_sizes = [3, 5]

    if batch_page < 1:
        batch_page = 1

    if batch_page_size not in allowed_batch_page_sizes:
        batch_page_size = 5

    batch_total_records = batch_query.count()

    batch_total_pages = (batch_total_records + batch_page_size - 1) // batch_page_size

    if batch_total_pages == 0:
        batch_total_pages = 1

    if batch_page > batch_total_pages:
        batch_page = batch_total_pages

    batch_offset = (batch_page - 1) * batch_page_size

    recent_batches = (
        batch_query
        .order_by(UploadBatch.id.desc())
        .offset(batch_offset)
        .limit(batch_page_size)
        .all()
    )

    recent_batch_items = []

    for batch in recent_batches:
        uploader = db.query(User).filter(User.id == batch.user_id).first()

        recent_batch_items.append(
            {
                "id": batch.id,
                "username": uploader.username if uploader else "未知上传方",
                "filename": batch.filename,
                "total_rows": batch.total_rows or 0,
                "success_rows": batch.success_rows or 0,
                "failed_rows": batch.failed_rows or 0,
                "acceptance_status": batch.acceptance_status or "待承接",
                "created_at": batch.created_at,
            }
        )    

    allowed_page_sizes = [3, 5, 10, 20]

    if page_size not in allowed_page_sizes:
        page_size = 10

    partners = (
        db.query(User)
        .filter(User.role == "partner")
        .order_by(User.id.desc())
        .all()
    )

    partner_options = [
        {
            "id": partner.id,
            "username": partner.username,
        }
        for partner in partners
    ]

    record_items, total_records, total_pages, page = build_business_record_items(
        db=db,
        user=user,
        partner_id=partner_id,
        keyword=keyword,
        start_date=start_date,
        end_date=end_date,
        review_status=review_status,
        acceptance_filter=acceptance_status,
        page=page,
        page_size=page_size,
        use_pagination=True,
    )

    db.close()

    return templates.TemplateResponse(
        "business_records.html",
        add_base_context(request, {
            "request": request,
            "active_page": "business_records",
            "partners": partner_options,
            "records": record_items,
            "recent_batches": recent_batch_items,

            "batch_page": batch_page,
            "batch_page_size": batch_page_size,
            "batch_total_records": batch_total_records,
            "batch_total_pages": batch_total_pages,
            "allowed_batch_page_sizes": allowed_batch_page_sizes,

            "partner_id": partner_id,
            "keyword": keyword,
            "start_date": start_date,
            "end_date": end_date,
            "review_status": review_status,
            "acceptance_status": acceptance_status,
            "page": page,
            "page_size": page_size,
            "total_records": total_records,
            "total_pages": total_pages,
            "allowed_page_sizes": allowed_page_sizes,
        }),
    )


@app.post("/upload-batches/{batch_id}/accept")
def accept_upload_batch(
    request: Request,
    batch_id: int,
    return_url: str = Form("/business-records"),
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()

    batch = db.query(UploadBatch).filter(UploadBatch.id == batch_id).first()

    if not batch:
        db.close()
        return RedirectResponse(url="/business-records", status_code=302)

    batch.acceptance_status = "已承接"

    db.commit()
    db.close()

    if not return_url.startswith("/") or return_url.startswith("//"):
        return_url = "/business-records"

    return RedirectResponse(url=return_url, status_code=302)


@app.post("/upload-batches/{batch_id}/reject")
def reject_upload_batch(
    request: Request,
    batch_id: int,
    return_url: str = Form("/business-records"),
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()

    batch = db.query(UploadBatch).filter(UploadBatch.id == batch_id).first()

    if not batch:
        db.close()
        return RedirectResponse(url="/business-records", status_code=302)

    batch.acceptance_status = "已拒绝"

    db.commit()
    db.close()

    if not return_url.startswith("/") or return_url.startswith("//"):
        return_url = "/business-records"

    return RedirectResponse(url=return_url, status_code=302)


@app.get("/business-records/export")
def export_business_records(
    request: Request,
    partner_id: int = Query(0),
    keyword: str = Query(""),
    start_date: str = Query(""),
    end_date: str = Query(""),
    review_status: str = Query("全部"),
    acceptance_status: str = Query("全部"),
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    db = SessionLocal()

    record_items, total_records, total_pages, page = build_business_record_items(
        db=db,
        user=user,
        partner_id=partner_id,
        keyword=keyword,
        start_date=start_date,
        end_date=end_date,
        review_status=review_status,
        acceptance_filter=acceptance_status,
        use_pagination=False,
    )

    export_rows = []

    for item in record_items:
        export_rows.append(
            {
                "业务单号": item["business_no"],
                "上传方": item["uploader_username"],
                "姓名": item["name"],
                "手机号": item["phone"],
                "车牌号": item["plate_number"],
                "积分金额": item["points_amount"],
                "银行卡号": item["bank_card"],
                "最新审核状态": item["latest_review_status"],
                "最新匹配状态": item["latest_match_status"],
                "已通过凭证金额": item["approved_voucher_amount"],
                "剩余金额": item["remaining_amount"],
                "导入时间": item["created_at"],
            }
        )

    total_points_amount = money2(sum(item["points_amount"] or 0 for item in record_items))
    total_approved_voucher_amount = money2(sum(item["approved_voucher_amount"] or 0 for item in record_items))
    total_remaining_amount = money2(sum(item["remaining_amount"] or 0 for item in record_items))

    partner_name = "全部上传方"

    if user.role == "partner":
        partner_name = user.username
    else:
        if partner_id != 0:
            selected_partner = db.query(User).filter(User.id == partner_id).first()
            partner_name = selected_partner.username if selected_partner else "未知上传方"

    df = pd.DataFrame(export_rows)

    summary_rows = [
        {"项目": "报表名称", "内容": "业务数据管理导出"},
        {"项目": "导出时间", "内容": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"项目": "上传方范围", "内容": partner_name},
        {"项目": "关键词", "内容": keyword or "全部"},
        {"项目": "开始日期", "内容": start_date or "不限"},
        {"项目": "结束日期", "内容": end_date or "不限"},
        {"项目": "审核状态", "内容": review_status or "全部"},
        {"项目": "承接状态", "内容": acceptance_status or "全部"},
        {"项目": "导出数据条数", "内容": total_records},
        {"项目": "积分金额合计", "内容": money2(total_points_amount)},
        {"项目": "已通过凭证金额合计", "内容": money2(total_approved_voucher_amount)},
        {"项目": "剩余金额合计", "内容": money2(total_remaining_amount)},
    ]

    summary_df = pd.DataFrame(summary_rows)

    os.makedirs("exports", exist_ok=True)

    filename = f"business_records_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = os.path.join("exports", filename)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="汇总说明")
        df.to_excel(writer, index=False, sheet_name="业务数据")

        summary_sheet = writer.sheets["汇总说明"]
        detail_sheet = writer.sheets["业务数据"]

        # 汇总说明页美化
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)

        summary_sheet.column_dimensions["A"].width = 24
        summary_sheet.column_dimensions["B"].width = 40

        for row in summary_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")

        # 业务数据页表头美化
        for cell in detail_sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        detail_sheet.freeze_panes = "A2"

        # 手机号、银行卡号、业务单号按文本格式处理
        text_columns = ["A", "D", "G"]

        for column_letter in text_columns:
            for cell in detail_sheet[column_letter]:
                cell.number_format = "@"

        # 金额列保留两位小数
        amount_columns = ["F", "J", "K"]

        for column_letter in amount_columns:
            for cell in detail_sheet[column_letter][1:]:
                cell.number_format = "0.00"

        # 自动列宽
        for worksheet in [summary_sheet, detail_sheet]:
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    value = cell.value
                    if value is None:
                        continue

                    max_length = max(max_length, len(str(value)))

                worksheet.column_dimensions[column_letter].width = min(max_length + 4, 45)

    db.close()

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/business-records/{record_id}", response_class=HTMLResponse)
def business_record_detail_page(
    request: Request,
    record_id: int,
    return_url: str = Query("/business-records"),
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    db = SessionLocal()

    record = (
        db.query(BusinessRecord)
        .filter(BusinessRecord.id == record_id)
        .first()
    )

    if not record:
        db.close()
        return RedirectResponse(url="/business-records", status_code=302)

    # 权限隔离：上传方只能看自己的业务数据
    if user.role == "partner" and record.user_id != user.id:
        db.close()
        return RedirectResponse(url="/dashboard", status_code=302)

    uploader = db.query(User).filter(User.id == record.user_id).first()

    batch = None
    if record.batch_id:
        batch = db.query(UploadBatch).filter(UploadBatch.id == record.batch_id).first()

    acceptance_status = batch.acceptance_status if batch and batch.acceptance_status else "待承接"

    reviews = (
        db.query(MatchReview)
        .filter(MatchReview.business_record_id == record.id)
        .order_by(MatchReview.id.desc())
        .all()
    )

    review_status_order = {
        "已通过": 0,
        "待审核": 1,
        "已驳回": 2,
    }

    reviews = sorted(
        reviews,
        key=lambda review: (
            review_status_order.get(review.review_status, 9),
            -review.id,
        ),
    )

    voucher_items = []
    approved_voucher_amount = 0.0

    for review in reviews:
        voucher = (
            db.query(VoucherRecord)
            .filter(VoucherRecord.id == review.voucher_id)
            .first()
        )

        if not voucher:
            continue

        voucher_url = ""

        if voucher.file_path:
            normalized_path = voucher.file_path.replace("\\", "/")

            if "/uploads/" in normalized_path:
                relative_path = normalized_path.split("/uploads/", 1)[1]
                voucher_url = "/uploads/" + relative_path
            elif normalized_path.startswith("uploads/"):
                voucher_url = "/" + normalized_path
            else:
                voucher_url = "/uploads/" + os.path.basename(normalized_path)
        elif voucher.filename:
            voucher_url = "/uploads/vouchers/" + voucher.filename

        voucher_amount = money2(voucher.voucher_amount or 0)

        if review.review_status == "已通过":
            approved_voucher_amount += voucher_amount

        ocr_text = voucher.ocr_text or ""
        ocr_excerpt = ocr_text[:120] + "..." if len(ocr_text) > 120 else ocr_text

        voucher_items.append(
            {
                "review_id": review.id,
                "voucher_id": voucher.id,
                "filename": voucher.filename,
                "voucher_amount": voucher_amount,
                "match_status": review.match_status,
                "review_status": review.review_status,
                "score": review.score,
                "name_match": review.name_match,
                "bank_match": review.bank_match,
                "amount_match": review.amount_match,
                "created_at": review.created_at,
                "voucher_url": voucher_url,
                "ocr_excerpt": ocr_excerpt,
            }
        )

    business_amount = money2(record.points_amount or 0)
    approved_voucher_amount = money2(approved_voucher_amount)
    raw_remaining_amount = money2(business_amount - approved_voucher_amount)

    remaining_amount = raw_remaining_amount
    if remaining_amount < 0:
        remaining_amount = 0.0

    overpaid_amount = 0.0
    if approved_voucher_amount > business_amount:
        overpaid_amount = money2(approved_voucher_amount - business_amount)

    if approved_voucher_amount <= 0:
        payment_status = "未付款"
    elif approved_voucher_amount < business_amount:
        payment_status = "部分付款"
    elif approved_voucher_amount == business_amount:
        payment_status = "已足额付款"
    else:
        payment_status = "超额付款"

    detail = {
        "id": record.id,
        "business_no": record.display_business_no,
        "uploader_username": uploader.username if uploader else "未知上传方",
        "batch_id": record.batch_id,
        "batch_filename": batch.filename if batch else "-",
        "acceptance_status": acceptance_status,
        "name": record.name,
        "phone": record.phone,
        "plate_number": record.plate_number,
        "points_amount": business_amount,
        "bank_card": record.bank_card,
        "record_service_rate": record.record_service_rate or 0,
        "record_upstream_cost_rate": record.record_upstream_cost_rate or 0,
        "created_at": record.created_at,
        "approved_voucher_amount": approved_voucher_amount,
        "remaining_amount": money2(remaining_amount),
        "overpaid_amount": overpaid_amount,
        "payment_status": payment_status,
        "voucher_count": len(voucher_items),
    }

    if not return_url.startswith("/") or return_url.startswith("//"):
        return_url = "/business-records"

    db.close()

    return templates.TemplateResponse(
        "business_record_detail.html",
        add_base_context(request, {
            "request": request,
            "active_page": "business_records",
            "detail": detail,
            "voucher_items": voucher_items,
            "return_url": return_url,
        }),
    )


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "error": None,
        },
    )


@app.post("/login", response_class=HTMLResponse)
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):
    db = SessionLocal()
    user = db.query(User).filter(User.username == username).first()
    db.close()

    if not user or not verify_password(password, user.password_hash):
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "用户名或密码错误",
            },
        )

    response = RedirectResponse(url="/dashboard", status_code=302)
    response.set_cookie(
        key="user_id",
        value=str(user.id),
        httponly=True,
        max_age=60 * 60 * 8,
    )
    return response


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    db = SessionLocal()
    
    pending_accept_batches = 0
    pending_voucher_batches = 0
    pending_review_records = 0
    today_finished_reviews = 0


    pending_accept_batches = (
        db.query(UploadBatch)
        .filter(UploadBatch.acceptance_status == "待承接")
        .count()
    )

    all_reviews = db.query(MatchReview).order_by(MatchReview.id.desc()).all()

    latest_reviews = []
    seen_business_record_ids = set()

    for review in all_reviews:
        if review.business_record_id in seen_business_record_ids:
            continue

        seen_business_record_ids.add(review.business_record_id)
        latest_reviews.append(review)

    pending_review_records = len([
        review for review in latest_reviews
        if review.review_status == "待审核"
    ])

    today_finished_reviews = len([
        review for review in latest_reviews
        if review.review_status in ["已通过", "已驳回"]
    ])

    pending_voucher_batch_ids = set()

    for review in latest_reviews:
        if review.review_status != "待审核":
            continue

        voucher = (
            db.query(VoucherRecord)
            .filter(VoucherRecord.id == review.voucher_id)
            .first()
        )

        if voucher and voucher.batch_id:
            pending_voucher_batch_ids.add(voucher.batch_id)

    pending_voucher_batches = len(pending_voucher_batch_ids)

    context = {
        "request": request,
        "username": user.username,
        "role": user.role,
        "active_page": "dashboard",
        "page_title": "后台首页",
        "pending_accept_batches": pending_accept_batches,
        "pending_voucher_batches": pending_voucher_batches,
        "pending_review_records": pending_review_records,
        "today_finished_reviews": today_finished_reviews,
    }

    db.close()

    return templates.TemplateResponse("dashboard.html", context)

RATE_MODE_LABELS = {
    EXTERNAL_MODE: "外扣",
    INTERNAL_MODE: "内扣",
}


def validate_partner_rate_config(
    rate: float,
    mode: str,
    field_label: str,
) -> str:
    """
    验证上传方的费率和计算方式配置。

    外扣：
        费率允许 0～100。

    内扣：
        费率允许 0～小于 100。
        因为 100% 会导致除数为 0。
    """
    normalized_mode = (
        mode or ""
    ).strip().lower()

    if normalized_mode not in RATE_MODE_LABELS:
        raise ValueError(
            f"{field_label}计算方式必须选择外扣或内扣"
        )

    if rate < 0 or rate > 100:
        raise ValueError(
            f"{field_label}必须在 0 到 100 之间"
        )

    if (
        normalized_mode == INTERNAL_MODE
        and rate >= 100
    ):
        raise ValueError(
            f"{field_label}采用内扣时必须小于 100"
        )

    return normalized_mode

@app.get("/partners", response_class=HTMLResponse)
def partners_page(
    request: Request,
    edit_id: int = Query(0),
    partner_page: int = Query(1),
    partner_page_size: int = Query(5),
    partner_keyword: str = Query(""),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()
    allowed_partner_page_sizes = [5, 10, 20]

    if partner_page_size not in allowed_partner_page_sizes:
        partner_page_size = 5

    if partner_page < 1:
        partner_page = 1

    partner_keyword = partner_keyword.strip()

    partner_query = db.query(User).filter(User.role == "partner")

    if partner_keyword:
        partner_query = partner_query.filter(User.username.contains(partner_keyword))

    partner_total = partner_query.count()
    partner_total_pages = max((partner_total + partner_page_size - 1) // partner_page_size, 1)

    if partner_page > partner_total_pages:
        partner_page = partner_total_pages

    partners = (
        partner_query
        .order_by(User.id.desc())
        .offset((partner_page - 1) * partner_page_size)
        .limit(partner_page_size)
        .all()
    )

    edit_partner = None
    if edit_id:
        edit_partner = (
            db.query(User)
            .filter(User.id == edit_id)
            .filter(User.role == "partner")
            .first()
        )

    db.close()

    return templates.TemplateResponse(
        "partners.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "partners",
            "partners": partners,
            "edit_partner": edit_partner,
            "edit_id": edit_id,
            "partner_page": partner_page,
            "partner_page_size": partner_page_size,
            "partner_total": partner_total,
            "partner_total_pages": partner_total_pages,
            "allowed_partner_page_sizes": allowed_partner_page_sizes,
            "partner_keyword": partner_keyword,
            "message": None,
            "error": None,
        },
    )


@app.post("/partners", response_class=HTMLResponse)
def create_partner(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    service_rate: float = Form(...),
    service_rate_mode: str = Form(...),
    upstream_cost_rate: float = Form(...),
    upstream_cost_rate_mode: str = Form(...),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()

    existing_user = db.query(User).filter(User.username == username).first()
    if existing_user:
        partners = db.query(User).filter(User.role == "partner").order_by(User.id.desc()).all()
        db.close()
        return templates.TemplateResponse(
            "partners.html",
            {
                "request": request,
                "username": user.username,
                "role": user.role,
                "topbar_username": user.username,
                "topbar_role": user.role,
                "active_page": "partners",
                "partners": partners,
                "message": None,
                "error": "该账号已存在，请换一个账号名",
            },
        )

    try:
        service_rate_mode = (
            validate_partner_rate_config(
                service_rate,
                service_rate_mode,
                "下游服务费率",
            )
        )

        upstream_cost_rate_mode = (
            validate_partner_rate_config(
                upstream_cost_rate,
                upstream_cost_rate_mode,
                "上游成本费率",
            )
        )

    except ValueError as exc:
        partners = (
            db.query(User)
            .filter(User.role == "partner")
            .order_by(User.id.desc())
            .all()
        )

        db.close()

        return templates.TemplateResponse(
            "partners.html",
            {
                "request": request,
                "username": user.username,
                "role": user.role,
                "topbar_username": user.username,
                "topbar_role": user.role,
                "active_page": "partners",
                "partners": partners,
                "message": None,
                "error": str(exc),
            },
        )

    new_partner = User(
        username=username,
        password_hash=get_password_hash(password),
        role="partner",
        service_rate=service_rate,
        service_rate_mode=service_rate_mode,
        upstream_cost_rate=upstream_cost_rate,
        upstream_cost_rate_mode=(
            upstream_cost_rate_mode
        ),
    )

    db.add(new_partner)
    db.commit()

    partners = db.query(User).filter(User.role == "partner").order_by(User.id.desc()).all()
    db.close()

    service_mode_label = RATE_MODE_LABELS[
        service_rate_mode
    ]

    upstream_mode_label = RATE_MODE_LABELS[
        upstream_cost_rate_mode
    ]

    return templates.TemplateResponse(
        "partners.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "partners",
            "partners": partners,
            "message": (
                f"上传方账号 {username} 创建成功，"
                f"下游服务费率为 {service_rate}%"
                f"（{service_mode_label}），"
                f"上游成本费率为 {upstream_cost_rate}%"
                f"（{upstream_mode_label}）"
            ),
            "error": None,
        },
    )


@app.get("/partners/{partner_id}/edit", response_class=HTMLResponse)
def edit_partner_page(request: Request, partner_id: int):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()

    partner = (
        db.query(User)
        .filter(User.id == partner_id)
        .filter(User.role == "partner")
        .first()
    )

    if not partner:
        db.close()
        return RedirectResponse(url="/partners", status_code=302)

    db.close()

    return templates.TemplateResponse(
        "edit_partner.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "partners",
            "partner": partner,
            "message": None,
            "error": None,
        },
    )


@app.post("/partners/{partner_id}/edit", response_class=HTMLResponse)
def edit_partner_submit(
    request: Request,
    partner_id: int,
    username: str = Form(...),
    service_rate: float = Form(...),
    service_rate_mode: str = Form(...),
    upstream_cost_rate: float = Form(...),
    upstream_cost_rate_mode: str = Form(...),
    partner_page: int = Form(1),
    partner_page_size: int = Form(5),
    partner_keyword: str = Form(""),
):
    user = get_current_user(request)

    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()

    partner = (
        db.query(User)
        .filter(User.id == partner_id)
        .filter(User.role == "partner")
        .first()
    )

    if not partner:
        db.close()
        return RedirectResponse(url="/partners", status_code=302)

    username = username.strip()

    if not username:
        db.close()
        return templates.TemplateResponse(
            "edit_partner.html",
            {
                "request": request,
                "username": user.username,
                "role": user.role,
                "topbar_username": user.username,
                "topbar_role": user.role,
                "active_page": "partners",
                "partner": partner,
                "message": None,
                "error": "上传方账号不能为空",
            },
        )

    existing_user = (
        db.query(User)
        .filter(User.username == username)
        .filter(User.id != partner.id)
        .first()
    )

    if existing_user:
        db.close()
        return templates.TemplateResponse(
            "edit_partner.html",
            {
                "request": request,
                "username": user.username,
                "role": user.role,
                "topbar_username": user.username,
                "topbar_role": user.role,
                "active_page": "partners",
                "partner": partner,
                "message": None,
                "error": "该账号名已存在，请换一个账号名",
            },
        )

    try:
        service_rate_mode = (
            validate_partner_rate_config(
                service_rate,
                service_rate_mode,
                "下游服务费率",
            )
        )

        upstream_cost_rate_mode = (
            validate_partner_rate_config(
                upstream_cost_rate,
                upstream_cost_rate_mode,
                "上游成本费率",
            )
        )

    except ValueError as exc:
        db.close()

        return templates.TemplateResponse(
            "edit_partner.html",
            {
                "request": request,
                "username": user.username,
                "role": user.role,
                "topbar_username": user.username,
                "topbar_role": user.role,
                "active_page": "partners",
                "partner": partner,
                "message": None,
                "error": str(exc),
            },
        )

    partner.username = username
    partner.service_rate = service_rate
    partner.service_rate_mode = service_rate_mode
    partner.upstream_cost_rate = upstream_cost_rate
    partner.upstream_cost_rate_mode = (
        upstream_cost_rate_mode
    )

    db.commit()
    db.refresh(partner)
    db.commit()
    db.refresh(partner)

    allowed_partner_page_sizes = [5, 10, 20]

    if partner_page_size not in allowed_partner_page_sizes:
        partner_page_size = 5

    if partner_page < 1:
        partner_page = 1


    partner_keyword = partner_keyword.strip()

    partner_query = db.query(User).filter(User.role == "partner")

    if partner_keyword:
        partner_query = partner_query.filter(User.username.contains(partner_keyword))

    partner_total = partner_query.count()
    partner_total_pages = max((partner_total + partner_page_size - 1) // partner_page_size, 1)

    if partner_page > partner_total_pages:
        partner_page = partner_total_pages

    partners = (
        partner_query
        .order_by(User.id.desc())
        .offset((partner_page - 1) * partner_page_size)
        .limit(partner_page_size)
        .all()
    )

    db.close()

    return templates.TemplateResponse(
        "partners.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "partners",
            "partners": partners,
            "edit_partner": partner,
            "edit_id": partner.id,
            "partner_page": partner_page,
            "partner_page_size": partner_page_size,
            "partner_total": partner_total,
            "partner_total_pages": partner_total_pages,
            "allowed_partner_page_sizes": allowed_partner_page_sizes,
            "partner_keyword": partner_keyword,
            "message": "上传方账号修改成功。新费率只会影响之后新上传的数据，历史业务数据继续使用原费率快照。",
            "error": None,
        },
    )


@app.get("/upload-excel", response_class=HTMLResponse)
def upload_excel_page(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse(
        "upload_excel.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "upload_excel",
            "message": None,
            "errors": None,
        },
    )


@app.post("/upload-excel", response_class=HTMLResponse)
async def upload_excel_submit(
    request: Request,
    file: UploadFile = File(...),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if not file.filename.endswith((".xlsx", ".xls")):
        return templates.TemplateResponse(
            "upload_excel.html",
            {
                "request": request,
                "username": user.username,
                "role": user.role,
                "message": None,
                "errors": ["只允许上传 Excel 文件：.xlsx 或 .xls"],
            },
        )

    os.makedirs("uploads/excel", exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_filename = f"{user.id}_{timestamp}_{file.filename}"
    file_path = os.path.join("uploads", "excel", safe_filename)

    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    records, errors = parse_business_excel(file_path)

    db = SessionLocal()

    batch = UploadBatch(
        user_id=user.id,
        filename=file.filename,
        total_rows=len(records) + len(errors),
        success_rows=len(records),
        failed_rows=len(errors),
        acceptance_status="待承接",
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)

    reserved_public_business_nos = set()

    for row_index, item in enumerate(records, start=1):
        batch_date = batch.created_at.strftime("%Y%m%d")
        business_no = f"BR{batch_date}B{batch.id:06d}R{row_index:06d}"

        business_record = BusinessRecord(
            user_id=user.id,
            batch_id=batch.id,
            business_no=business_no,
            public_business_no=generate_unique_public_business_no(
                db,
                reserved_public_business_nos,
            ),
            name=item["name"],
            phone=item["phone"],
            plate_number=item["plate_number"],
            points_amount=item["points_amount"],
            bank_card=item["bank_card"],
            record_service_rate=user.service_rate or 0,
            record_upstream_cost_rate=user.upstream_cost_rate or 0,
            record_service_rate_mode=(
                user.service_rate_mode
                if user.service_rate_mode in (
                    "external",
                    "internal",
                )
                else "external"
            ),
            record_upstream_cost_rate_mode=(
                user.upstream_cost_rate_mode
                if user.upstream_cost_rate_mode in (
                    "external",
                    "internal",
                )
                else "external"
            ),
        )
        db.add(business_record)

    db.commit()
    db.close()

    message = f"上传成功：共读取 {len(records) + len(errors)} 行，成功导入 {len(records)} 行，失败 {len(errors)} 行。"

    return templates.TemplateResponse(
        "upload_excel.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "message": message,
            "errors": errors,
        },
    )


def build_accepted_business_batch_options(db, partner_id: int):
    if not partner_id or partner_id == 0:
        return []

    batches = (
        db.query(UploadBatch)
        .filter(UploadBatch.user_id == partner_id)
        .filter(UploadBatch.acceptance_status == ACCEPTED_BATCH_STATUS)
        .order_by(UploadBatch.id.desc())
        .all()
    )

    options = []

    for batch in batches:
        created_text = "-"
        if batch.created_at:
            created_text = batch.created_at.strftime("%Y-%m-%d %H:%M")

        options.append(
            {
                "id": batch.id,
                "label": f"批次 {batch.id}｜{batch.filename}｜成功 {batch.success_rows or 0} 条｜{created_text}",
            }
        )

    return options

def build_voucher_matching_records(db, partner_id: int, selected_business_batch_id: int):
    # OCR 匹配源只使用已承接批次下的业务数据
    records_query = apply_accepted_batch_filter(db.query(BusinessRecord))

    if partner_id != 0:
        records_query = records_query.filter(BusinessRecord.user_id == partner_id)

    # 如果管理员选择了某一份已承接业务清单，则只在该清单对应的业务数据里匹配
    if selected_business_batch_id != 0:
        selected_batch_query = (
            db.query(UploadBatch)
            .filter(UploadBatch.id == selected_business_batch_id)
            .filter(UploadBatch.acceptance_status == ACCEPTED_BATCH_STATUS)
        )

        if partner_id != 0:
            selected_batch_query = selected_batch_query.filter(
                UploadBatch.user_id == partner_id
            )

        selected_batch = selected_batch_query.first()

        if selected_batch:
            records_query = records_query.filter(
                BusinessRecord.batch_id == selected_business_batch_id
            )
        else:
            records_query = records_query.filter(BusinessRecord.id == -1)

    return records_query.all()


def create_match_reviews_for_voucher(db, voucher_record, records):
    raw_match_results = match_ocr_with_records(
        voucher_record.ocr_text or "",
        records,
        voucher_amount=voucher_record.voucher_amount or 0,
    )

    created_review_count = 0
    skipped_review_count = 0

    for item in raw_match_results:
        record = item["record"]

        existing_review = (
            db.query(MatchReview)
            .filter(MatchReview.voucher_id == voucher_record.id)
            .filter(MatchReview.business_record_id == record.id)
            .first()
        )

        if existing_review:
            skipped_review_count += 1
            continue

        review = MatchReview(
            voucher_id=voucher_record.id,
            business_record_id=record.id,
            match_status=item["status"],
            name_match=item.get("name_detail", "未知"),
            bank_match=item.get(
                "bank_detail",
                "是" if item["bank_match"] else "否",
            ),
            amount_match=(
                "是"
                if item["amount_match"]
                else item.get("partial_amount_detail", "否")
            ),
            score=item["score"],
            review_status="待审核",
        )

        db.add(review)
        created_review_count += 1

    db.commit()

    return created_review_count, skipped_review_count

def build_voucher_upload_batch_items(db, page: int = 1, page_size: int = 5):
    if page < 1:
        page = 1

    allowed_page_sizes = [5, 10, 20]
    if page_size not in allowed_page_sizes:
        page_size = 5

    total_batches = db.query(VoucherUploadBatch).count()

    total_pages = (total_batches + page_size - 1) // page_size
    if total_pages < 1:
        total_pages = 1

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * page_size

    batches = (
        db.query(VoucherUploadBatch)
        .order_by(VoucherUploadBatch.id.desc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    batch_items = []

    for batch in batches:
        partner_name = "全部上传方"

        if batch.partner_id and batch.partner_id != 0:
            partner = db.query(User).filter(User.id == batch.partner_id).first()
            if partner:
                partner_name = partner.username

        uploader = db.query(User).filter(User.id == batch.uploader_id).first()

        linked_voucher_count = (
            db.query(VoucherRecord)
            .filter(VoucherRecord.batch_id == batch.id)
            .count()
        )

        total_review_count = (
            db.query(MatchReview)
            .join(VoucherRecord, MatchReview.voucher_id == VoucherRecord.id)
            .filter(VoucherRecord.batch_id == batch.id)
            .count()
        )

        pending_review_count = (
            db.query(MatchReview)
            .join(VoucherRecord, MatchReview.voucher_id == VoucherRecord.id)
            .filter(VoucherRecord.batch_id == batch.id)
            .filter(MatchReview.review_status == "待审核")
            .count()
        )

        processed_review_count = (
            db.query(MatchReview)
            .join(VoucherRecord, MatchReview.voucher_id == VoucherRecord.id)
            .filter(VoucherRecord.batch_id == batch.id)
            .filter(MatchReview.review_status.in_(["已通过", "已驳回"]))
            .count()
        )

        if total_review_count == 0:
            review_action_text = "无匹配记录"
            review_action_class = "muted"
        elif pending_review_count > 0 and processed_review_count > 0:
            review_action_text = "继续审核"
            review_action_class = "primary"
        elif pending_review_count > 0:
            review_action_text = "去审核"
            review_action_class = "primary"
        else:
            review_action_text = "查看结果"
            review_action_class = "secondary"

        batch_items.append(
            {
                "id": batch.id,
                "uploader_username": uploader.username if uploader else "未知用户",
                "partner_name": partner_name,
                "total_files": batch.total_files or 0,
                "success_files": batch.success_files or 0,
                "duplicate_files": batch.duplicate_files or 0,
                "failed_files": batch.failed_files or 0,
                "linked_voucher_count": linked_voucher_count,
                "total_created_reviews": total_review_count,
                "pending_review_count": pending_review_count,
                "processed_review_count": processed_review_count,
                "review_action_text": review_action_text,
                "review_action_class": review_action_class,
                "created_at": batch.created_at,
            }
        )

    pagination = {
        "voucher_page": page,
        "voucher_page_size": page_size,
        "voucher_total_batches": total_batches,
        "voucher_total_pages": total_pages,
    }

    return batch_items, pagination


@app.get("/upload-voucher", response_class=HTMLResponse)
def upload_voucher_page(
    request: Request,
    voucher_page: int = Query(1),
    voucher_page_size: int = Query(5),
    partner_id: int = Query(0),
    selected_business_batch_id: int = Query(0),
):
    user = get_current_user(request)

    if not user or user.role != "admin":
        return RedirectResponse(url="/login", status_code=302)

    db = SessionLocal()

    partners = (
        db.query(User)
        .filter(User.role == "partner")
        .order_by(User.id.desc())
        .all()
    )

    partner_options = [
        {
            "id": partner.id,
            "username": partner.username,
        }
        for partner in partners
    ]

    voucher_batches, voucher_pagination = build_voucher_upload_batch_items(
        db,
        page=voucher_page,
        page_size=voucher_page_size,
    )

    accepted_business_batches = build_accepted_business_batch_options(
        db,
        partner_id,
    )

    db.close()

    return templates.TemplateResponse(
        "upload_voucher.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "upload_voucher",
            "partners": partner_options,
            "selected_partner_id": partner_id,
            "selected_business_batch_id": selected_business_batch_id,
            "accepted_business_batches": accepted_business_batches,
            "voucher_batches": voucher_batches,
            "voucher_page": voucher_pagination["voucher_page"],
            "voucher_page_size": voucher_pagination["voucher_page_size"],
            "voucher_total_batches": voucher_pagination["voucher_total_batches"],
            "voucher_total_pages": voucher_pagination["voucher_total_pages"],
        },
    )


@app.post("/upload-voucher", response_class=HTMLResponse)
def upload_voucher_submit(
    request: Request,
    files: list[UploadFile] = File(...),
    partner_id: int = Form(0),
    selected_business_batch_id: int = Form(0),
):
    user = get_current_user(request)

    if not user or user.role != "admin":
        return RedirectResponse(url="/login", status_code=302)

    db = SessionLocal()

    partners = (
        db.query(User)
        .filter(User.role == "partner")
        .order_by(User.id.desc())
        .all()
    )

    partner_options = [
        {
            "id": partner.id,
            "username": partner.username,
        }
        for partner in partners
    ]

    accepted_business_batches = build_accepted_business_batch_options(
        db,
        partner_id,
    )

    os.makedirs("uploads/vouchers", exist_ok=True)

    batch_results = []
    total_created_reviews = 0

    success_files = 0
    duplicate_files = 0
    failed_files = 0

    voucher_batch = VoucherUploadBatch(
        uploader_id=user.id,
        partner_id=partner_id,
        total_files=len(files),
        success_files=0,
        duplicate_files=0,
        failed_files=0,
        total_created_reviews=0,
    )

    db.add(voucher_batch)
    db.commit()
    db.refresh(voucher_batch)

    for file in files:
        if not file or not file.filename:
            continue

        try:
            content = file.file.read()
            file_hash = hashlib.sha256(content).hexdigest()

            existing_voucher = (
                db.query(VoucherRecord)
                .filter(VoucherRecord.file_hash == file_hash)
                .first()
            )

            if existing_voucher:
                duplicate_files += 1

                # 如果重复凭证这次选择了已承接清单，则复用已有凭证重新匹配当前清单
                if selected_business_batch_id != 0:
                    records = build_voucher_matching_records(
                        db,
                        partner_id,
                        selected_business_batch_id,
                    )

                    created_review_count, skipped_review_count = create_match_reviews_for_voucher(
                        db,
                        existing_voucher,
                        records,
                    )

                    total_created_reviews += created_review_count

                    if created_review_count > 0:
                        message = (
                            f"复用凭证ID：{existing_voucher.id}，"
                            f"当前清单新增 {created_review_count} 条审核记录"
                        )
                    elif skipped_review_count > 0:
                        message = (
                            f"复用凭证ID：{existing_voucher.id}，"
                            f"当前清单已有审核记录，跳过 {skipped_review_count} 条"
                        )
                    else:
                        message = (
                            f"复用凭证ID：{existing_voucher.id}，"
                            f"当前清单内未找到匹配业务"
                        )

                    batch_results.append(
                        {
                            "filename": file.filename,
                            "status": "复用匹配",
                            "message": message,
                            "voucher_amount": existing_voucher.voucher_amount or 0,
                            "match_count": created_review_count,
                        }
                    )
                else:
                    batch_results.append(
                        {
                            "filename": file.filename,
                            "status": "重复跳过",
                            "message": f"凭证ID：{existing_voucher.id} 已存在；如需重匹配，请先选择已承接清单。",
                            "voucher_amount": existing_voucher.voucher_amount or 0,
                            "match_count": 0,
                        }
                    )

                continue

            safe_filename = file.filename.replace("\\", "_").replace("/", "_")
            saved_filename = (
                f"{user.id}_"
                f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_"
                f"{safe_filename}"
            )
            file_path = os.path.join("uploads", "vouchers", saved_filename)

            with open(file_path, "wb") as f:
                f.write(content)

            ocr_text = ocr_image(file_path)
            voucher_amount = extract_voucher_amount(ocr_text) or 0

            voucher_record = VoucherRecord(
                uploader_id=user.id,
                batch_id=voucher_batch.id,
                filename=file.filename,
                file_path=file_path,
                file_hash=file_hash,
                voucher_amount=voucher_amount,
                ocr_text=ocr_text,
            )

            db.add(voucher_record)
            db.commit()
            db.refresh(voucher_record)

            records = build_voucher_matching_records(
                db,
                partner_id,
                selected_business_batch_id,
            )

            created_review_count, skipped_review_count = create_match_reviews_for_voucher(
                db,
                voucher_record,
                records,
            )

            success_files += 1
            total_created_reviews += created_review_count

            if created_review_count > 0:
                result_status = "识别成功"
                result_message = f"新增 {created_review_count} 条审核记录"
            else:
                result_status = "未匹配"
                result_message = "当前清单内未找到匹配业务"

            batch_results.append(
                {
                    "filename": file.filename,
                    "status": result_status,
                    "voucher_amount": voucher_amount,
                    "match_count": created_review_count,
                    "message": result_message,
                }
            )

        except Exception as e:
            db.rollback()
            failed_files += 1

            batch_results.append(
                {
                    "filename": file.filename,
                    "status": "处理失败",
                    "message": str(e),
                    "voucher_amount": 0,
                    "match_count": 0,
                }
            )

    voucher_batch.success_files = success_files
    voucher_batch.duplicate_files = duplicate_files
    voucher_batch.failed_files = failed_files
    voucher_batch.total_created_reviews = total_created_reviews
    db.commit()

    voucher_batches, voucher_pagination = build_voucher_upload_batch_items(
        db,
        page=1,
        page_size=5,
    )

    db.close()

    return templates.TemplateResponse(
        "upload_voucher.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "upload_voucher",
            "partners": partner_options,
            "selected_partner_id": partner_id,
            "selected_business_batch_id": selected_business_batch_id,
            "accepted_business_batches": accepted_business_batches,
            "batch_results": batch_results,
            "total_files": len(files),
            "total_created_reviews": total_created_reviews,
            "voucher_batches": voucher_batches,
            "voucher_batches": voucher_batches,
            "voucher_page": voucher_pagination["voucher_page"],
            "voucher_page_size": voucher_pagination["voucher_page_size"],
            "voucher_total_batches": voucher_pagination["voucher_total_batches"],
            "voucher_total_pages": voucher_pagination["voucher_total_pages"],
        },
    )

@app.get("/match-reviews", response_class=HTMLResponse)
def match_reviews_page(
    request: Request,
    status_filter: str = Query("全部"),
    partner_id: int = Query(0),
    page: int = Query(1),
    page_size: int = Query(1),
    customer_name: str = Query(""),
    review_id: int = Query(0),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()

    partners = db.query(User).filter(User.role == "partner").order_by(User.id.desc()).all()

    if review_id > 0:
        reviews = (
            db.query(MatchReview)
            .filter(MatchReview.id == review_id)
            .order_by(MatchReview.id.desc())
            .all()
        )
    else:
        reviews = db.query(MatchReview).order_by(MatchReview.id.desc()).all()

    # 匹配结果审核页的展示单位应该是“审核记录 MatchReview”，不是“业务数据 BusinessRecord”。
    # 同一条业务可能关联多张凭证，其中既可能有已通过，也可能有已驳回、待审核、无需审核。
    # 如果按 business_record_id 去重，只保留最新一条，就会把较早的“已通过”历史审核记录隐藏掉。
    latest_reviews = reviews

    if status_filter in ["待审核", "已通过", "已驳回"]:
        latest_reviews = [
            review for review in latest_reviews
            if review.review_status == status_filter
        ]

    if partner_id != 0:
        filtered_reviews = []

        for review in latest_reviews:
            record = db.query(BusinessRecord).filter(
                BusinessRecord.id == review.business_record_id
            ).first()

            if record and record.user_id == partner_id:
                filtered_reviews.append(review)

        latest_reviews = filtered_reviews

    customer_name = customer_name.strip()

    if customer_name:
        filtered_reviews = []

        for review in latest_reviews:
            record = db.query(BusinessRecord).filter(
                BusinessRecord.id == review.business_record_id
            ).first()

            if record and customer_name in (record.name or ""):
                filtered_reviews.append(review)

        latest_reviews = filtered_reviews

    allowed_page_sizes = [1, 3, 5, 10, 20]

    if page_size not in allowed_page_sizes:
        page_size = 1

    total_reviews = len(latest_reviews)

    total_pages = (total_reviews + page_size - 1) // page_size

    if total_pages == 0:
        total_pages = 1

    if page < 1:
        page = 1

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * page_size

    page_reviews = latest_reviews[offset: offset + page_size]

    review_items = []

    for review in page_reviews:
        voucher = db.query(VoucherRecord).filter(VoucherRecord.id == review.voucher_id).first()
        record = db.query(BusinessRecord).filter(BusinessRecord.id == review.business_record_id).first()

        if voucher and record:
            uploader = db.query(User).filter(User.id == record.user_id).first()

            voucher_url = ""

            if voucher.file_path:
                normalized_path = voucher.file_path.replace("\\", "/")

                if "/uploads/" in normalized_path:
                    relative_path = normalized_path.split("/uploads/", 1)[1]
                    voucher_url = "/uploads/" + relative_path
                elif normalized_path.startswith("uploads/"):
                    voucher_url = "/" + normalized_path
                else:
                    voucher_url = "/uploads/" + os.path.basename(normalized_path)
            elif voucher.filename:
                voucher_url = "/uploads/vouchers/" + voucher.filename
            
            service_rate = record.record_service_rate if record.record_service_rate is not None else 0
            upstream_cost_rate = (
                record.record_upstream_cost_rate
                if record.record_upstream_cost_rate is not None
                else 0
            )
            points_amount = record.points_amount or 0

            service_rate_mode = (
                record.record_service_rate_mode
                if record.record_service_rate_mode
                in (
                    "external",
                    "internal",
                )
                else "external"
            )

            upstream_cost_rate_mode = (
                record.record_upstream_cost_rate_mode
                if record.record_upstream_cost_rate_mode
                in (
                    "external",
                    "internal",
                )
                else "external"
            )

            settlement_result = calculate_business_settlement(
                base_amount=points_amount,
                downstream_rate_percent=service_rate,
                downstream_mode=service_rate_mode,
                upstream_rate_percent=upstream_cost_rate,
                upstream_mode=upstream_cost_rate_mode,
            )

            receivable_fee = (
                settlement_result.downstream.fee_amount
            )

            payable_cost = (
                settlement_result.upstream.fee_amount
            )

            gross_profit = settlement_result.gross_profit

            approved_reviews_for_record = (
                db.query(MatchReview)
                .filter(MatchReview.business_record_id == record.id)
                .filter(MatchReview.review_status == "已通过")
                .all()
            )

            approved_voucher_amount = 0

            for approved_review in approved_reviews_for_record:
                approved_voucher = (
                    db.query(VoucherRecord)
                    .filter(VoucherRecord.id == approved_review.voucher_id)
                    .first()
                )

                if approved_voucher and approved_voucher.voucher_amount:
                    approved_voucher_amount += approved_voucher.voucher_amount

            business_amount = record.points_amount or 0
            remaining_amount = business_amount - approved_voucher_amount

            if remaining_amount < 0:
                remaining_amount = 0

            review_items.append(
                {
                    "review": review,
                    "voucher": voucher,
                    "record": record,
                    "voucher_url": voucher_url,
                    "uploader_username": uploader.username if uploader else "未知上传方",
                    "service_rate": service_rate,
                    "upstream_cost_rate": upstream_cost_rate,
                    "receivable_fee": float(receivable_fee),
                    "payable_cost": float(payable_cost),
                    "gross_profit": float(gross_profit),
                    "voucher_amount": voucher.voucher_amount or 0,
                    "approved_voucher_amount": round(approved_voucher_amount, 2),
                    "remaining_amount": round(remaining_amount, 2),
                }
            )

    db.close()

    return templates.TemplateResponse(
        "match_reviews.html",
        {
            "request": request,
            "username": user.username,
            "review_items": review_items,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "match_reviews",
            "reviews": review_items,
            "status_filter": status_filter,
            "partners": partners,
            "partner_id": partner_id,
            "page": page,
            "page_size": page_size,
            "customer_name": customer_name,
            "review_id": review_id,
            "total_reviews": total_reviews,
            "total_pages": total_pages,
            "allowed_page_sizes": allowed_page_sizes,
        },
    )

@app.post("/match-reviews/{review_id}/approve")
def approve_match_review(
    review_id: int, 
    request: Request,
    status_filter: str = Form("全部"),
    partner_id: int = Form(0),
    page: int = Form(1),
    page_size: int = Form(1),
    customer_name: str = Form(""),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()
    review = db.query(MatchReview).filter(MatchReview.id == review_id).first()

    if review:
        review.review_status = "已通过"
        db.commit()

    db.close()

    redirect_url = "/match-reviews?" + urlencode(
        {
            "status_filter": status_filter,
            "partner_id": partner_id,
            "page": page,
            "page_size": page_size,
            "customer_name": customer_name,
        }
    )   

    return RedirectResponse(url=redirect_url, status_code=302)


@app.post("/match-reviews/{review_id}/reject")
def reject_match_review(
    review_id: int, 
    request: Request,
    status_filter: str = Form("全部"),
    partner_id: int = Form(0),
    page: int = Form(1),
    page_size: int = Form(1),
    customer_name: str = Form(""),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    db = SessionLocal()
    review = db.query(MatchReview).filter(MatchReview.id == review_id).first()

    if review:
        review.review_status = "已驳回"
        db.commit()

    db.close()

    redirect_url = "/match-reviews?" + urlencode(
        {
            "status_filter": status_filter,
            "partner_id": partner_id,
            "page": page,
            "page_size": page_size,
            "customer_name": customer_name,
        }
    )

    return RedirectResponse(url=redirect_url, status_code=302)


@app.post("/match-reviews/{review_id}/reopen")
def reopen_match_review(
    review_id: int,
    request: Request,
    status_filter: str = Form("全部"),
    partner_id: int = Form(0),
    page: int = Form(1),
    page_size: int = Form(1),
    customer_name: str = Form(""),
):
    user = get_current_user(request)

    if not user or user.role != "admin":
        return RedirectResponse(url="/login", status_code=302)

    db = SessionLocal()

    review = db.query(MatchReview).filter(MatchReview.id == review_id).first()

    if review:
        review.review_status = "待审核"
        db.commit()

    db.close()

    redirect_url = "/match-reviews?" + urlencode(
        {
            "status_filter": status_filter,
            "partner_id": partner_id,
            "page": page,
            "page_size": page_size,
            "customer_name": customer_name,
        }
    )

    return RedirectResponse(url=redirect_url, status_code=302)


@app.post("/match-reviews/batch-review")
def batch_review_match_reviews(
    request: Request,
    review_ids: list[int] = Form([]),
    action: str = Form(...),
    status_filter: str = Form("全部"),
    partner_id: int = Form(0),
    page: int = Form(1),
    page_size: int = Form(1),
    customer_name: str = Form(""),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    if action not in ["approve", "reject"]:
        query_params = []

        if status_filter and status_filter != "全部":
            query_params.append(f"status_filter={status_filter}")

        if partner_id != 0:
            query_params.append(f"partner_id={partner_id}")

        redirect_url = "/match-reviews"

        if query_params:
            redirect_url = "/match-reviews?" + "&".join(query_params)

        return RedirectResponse(url=redirect_url, status_code=302)

    new_status = "已通过" if action == "approve" else "已驳回"

    db = SessionLocal()

    if review_ids:
        db.query(MatchReview).filter(MatchReview.id.in_(review_ids)).update(
            {MatchReview.review_status: new_status},
            synchronize_session=False,
        )
        db.commit()

    db.close()

    redirect_url = "/match-reviews?" + urlencode(
        {
            "status_filter": status_filter,
            "partner_id": partner_id,
            "page": page,
            "page_size": page_size,
            "customer_name": customer_name,
        }
    )

    return RedirectResponse(url=redirect_url, status_code=302)

@app.get("/upload-batches", response_class=HTMLResponse)
def upload_batches_page(
    request: Request,
    partner_id: int = Query(0),
    keyword: str = Query(""),
    start_date: str = Query(""),
    end_date: str = Query(""),
    review_status: str = Query("全部"),
    acceptance_status: str = Query("全部"),
    batch_page: int = 1,
    batch_page_size: int = 5,
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    db = SessionLocal()

    partners = []

    if user.role == "admin":
        partners = (
            db.query(User)
            .filter(User.role == "partner")
            .order_by(User.id.desc())
            .all()
        )

    allowed_acceptance_statuses = ["全部", "待承接", "已承接", "已拒绝"]

    if acceptance_status not in allowed_acceptance_statuses:
        acceptance_status = "全部"

    query = db.query(UploadBatch)

    # 权限隔离：
    # 管理员可以查看全部上传记录
    # 上传方只能查看自己的上传记录
    if user.role != "admin":
        query = query.filter(UploadBatch.user_id == user.id)
    else:
        if partner_id != 0:
            query = query.filter(UploadBatch.user_id == partner_id)

    if start_date:
        start_dt = datetime.combine(datetime.strptime(start_date, "%Y-%m-%d").date(), time.min)
        query = query.filter(UploadBatch.created_at >= start_dt)

    if end_date:
        end_dt = datetime.combine(datetime.strptime(end_date, "%Y-%m-%d").date(), time.max)
        query = query.filter(UploadBatch.created_at <= end_dt)

    if acceptance_status != "全部":
        query = query.filter(UploadBatch.acceptance_status == acceptance_status)

    allowed_batch_page_sizes = [5, 10, 20]

    if batch_page_size not in allowed_batch_page_sizes:
        batch_page_size = 5

    if batch_page < 1:
        batch_page = 1

    batch_total = query.count()
    batch_total_pages = max((batch_total + batch_page_size - 1) // batch_page_size, 1)

    if batch_page > batch_total_pages:
        batch_page = batch_total_pages

    batches = (
        query
        .order_by(UploadBatch.id.desc())
        .offset((batch_page - 1) * batch_page_size)
        .limit(batch_page_size)
        .all()
    )

    batch_items = []

    for batch in batches:
        uploader = db.query(User).filter(User.id == batch.user_id).first()

        batch_items.append(
            {
                "id": batch.id,
                "username": uploader.username if uploader else "未知用户",
                "filename": batch.filename,
                "total_rows": batch.total_rows,
                "success_rows": batch.success_rows,
                "failed_rows": batch.failed_rows,
                "acceptance_status": batch.acceptance_status or "待承接",
                "created_at": batch.created_at,
            }
        )

    db.close()

    return templates.TemplateResponse(
        "upload_batches.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "upload_batches",
            "batches": batch_items,
            "partners": partners,
            "partner_id": partner_id,
            "keyword": keyword,
            "start_date": start_date,
            "end_date": end_date,
            "review_status": review_status,
            "acceptance_status": acceptance_status,
            "batch_page": batch_page,
            "batch_page_size": batch_page_size,
            "batch_total": batch_total,
            "batch_total_pages": batch_total_pages,
            "allowed_batch_page_sizes": allowed_batch_page_sizes,
        },
    )

@app.get("/stats-dashboard", response_class=HTMLResponse)
def stats_dashboard_page(
    request: Request,
    partner_id: int = Query(0),
    start_date: str = Query(""),
    end_date: str = Query(""),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    stats = build_stats_data(
        partner_id=partner_id,
        start_date=start_date,
        end_date=end_date,
    )

    return templates.TemplateResponse(
        "stats_dashboard.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "stats_dashboard",
            **stats,
        },
    )

@app.get("/stats-dashboard/export")
def export_stats_dashboard(
    request: Request,
    partner_id: int = Query(0),
    start_date: str = Query(""),
    end_date: str = Query(""),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    stats = build_stats_data(
        partner_id=partner_id,
        start_date=start_date,
        end_date=end_date,
    )

    os.makedirs("exports", exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    export_path = os.path.join("exports", f"settlement_summary_{timestamp}.xlsx")

    selected_partner_name = "全部上传方"

    if partner_id != 0:
        for partner in stats["partners"]:
            if partner.id == partner_id:
                selected_partner_name = partner.username
                break

    export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    summary_rows = [
        {"指标": "报表名称", "数值": "结算汇总报表"},
        {"指标": "导出时间", "数值": export_time},
        {"指标": "上传方范围", "数值": selected_partner_name},
        {"指标": "开始日期", "数值": start_date if start_date else "全部"},
        {"指标": "结束日期", "数值": end_date if end_date else "全部"},
        {"指标": "", "数值": ""},
        {"指标": "业务数据总条数", "数值": stats["total_records"]},
        {"指标": "上传批次数", "数值": stats["total_batches"]},
        {"指标": "上传方账号数", "数值": stats["total_partners"]},
        {"指标": "待审核数量", "数值": stats["pending_reviews"]},
        {"指标": "已通过数量", "数值": stats["approved_reviews"]},
        {"指标": "已驳回数量", "数值": stats["rejected_reviews"]},
        {"指标": "总积分金额", "数值": stats["total_points"]},
        {"指标": "应收服务费合计", "数值": stats["total_receivable_fee"]},
        {"指标": "应付成本费合计", "数值": stats["total_payable_cost"]},
        {"指标": "毛利合计", "数值": stats["total_gross_profit"]},
        {"指标": "", "数值": ""},
        {"指标": "已通过结算条数", "数值": stats["approved_settlement_count"]},
        {"指标": "已通过结算金额", "数值": stats["approved_settlement_points"]},
        {"指标": "已通过应收服务费", "数值": stats["approved_settlement_receivable_fee"]},
        {"指标": "已通过应付成本费", "数值": stats["approved_settlement_payable_cost"]},
        {"指标": "已通过毛利", "数值": stats["approved_settlement_gross_profit"]},       
    ]

    detail_df = pd.DataFrame(stats["rows"])

    partner_summary_rows = []

    if not detail_df.empty:
        grouped = detail_df.groupby("上传方")

        for partner_name, group in grouped:
            total_points = group["积分金额"].sum()

            if "下游服务率" in group.columns:
                service_rate_column = "下游服务率"
            elif "下游费率" in group.columns:
                service_rate_column = "下游费率"
            else:
                service_rate_column = None

            if "上游成本率" in group.columns:
                upstream_cost_rate_column = "上游成本率"
            elif "上游成本费率" in group.columns:
                upstream_cost_rate_column = "上游成本费率"
            else:
                upstream_cost_rate_column = None

            if service_rate_column:
                service_rates = group[service_rate_column].dropna().unique()
                service_rate = service_rates[0] if len(service_rates) == 1 else "多版本费率"
            else:
                service_rate = "多版本费率"

            if upstream_cost_rate_column:
                upstream_cost_rates = group[upstream_cost_rate_column].dropna().unique()
                upstream_cost_rate = upstream_cost_rates[0] if len(upstream_cost_rates) == 1 else "多版本费率"
            else:
                upstream_cost_rate = "多版本费率"

            total_receivable_fee = group["应收服务费"].sum()
            total_payable_cost = group["应付成本费"].sum()
            total_gross_profit = group["毛利"].sum()

            approved_group = group[group["审核状态"] == "已通过"]

            approved_points = approved_group["积分金额"].sum()
            approved_receivable_fee = approved_group["应收服务费"].sum()
            approved_payable_cost = approved_group["应付成本费"].sum()
            approved_gross_profit = approved_group["毛利"].sum()

            partner_summary_rows.append(
                {
                    "上传方": partner_name,
                    "业务条数": len(group),
                    "总积分金额": round(total_points, 2),
                    "下游服务费率": service_rate,
                    "应收服务费合计": round(total_receivable_fee, 2),
                    "上游成本费率": upstream_cost_rate,
                    "应付成本费合计": round(total_payable_cost, 2),
                    "毛利合计": round(total_gross_profit, 2),
                    "已通过结算条数": len(approved_group),
                    "已通过结算金额": round(approved_points, 2),
                    "已通过应收服务费": round(approved_receivable_fee, 2),
                    "已通过应付成本费": round(approved_payable_cost, 2),
                    "已通过毛利": round(approved_gross_profit, 2),
                }
            )

    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="汇总", index=False)
        pd.DataFrame(partner_summary_rows).to_excel(writer, sheet_name="上传方汇总", index=False)
        detail_df.to_excel(writer, sheet_name="明细", index=False)
        
        format_excel_file(writer)

    return FileResponse(
        export_path,
        filename=f"结算汇总_{timestamp}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/my-settlement", response_class=HTMLResponse)
def my_settlement_page(
    request: Request,
    start_date: str = Query(""),
    end_date: str = Query(""),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "partner":
        return RedirectResponse(url="/dashboard", status_code=302)

    stats = build_stats_data(
        partner_id=user.id,
        start_date=start_date,
        end_date=end_date,
    )

    return templates.TemplateResponse(
        "my_settlement.html",
        {
            "request": request,
            "username": user.username,
            "role": user.role,
            "topbar_username": user.username,
            "topbar_role": user.role,
            "active_page": "my_settlement",
            "start_date": start_date,
            "end_date": end_date,
            "total_records": stats["total_records"],
            "total_points": stats["total_points"],
            "approved_settlement_count": stats["approved_settlement_count"],
            "approved_settlement_points": stats["approved_settlement_points"],
            "settlement_service_fee": stats["approved_settlement_receivable_fee"],
        },
    )

@app.get("/my-settlement/export")
def export_my_settlement(
    request: Request,
    start_date: str = Query(""),
    end_date: str = Query(""),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    if user.role != "partner":
        return RedirectResponse(url="/dashboard", status_code=302)

    stats = build_stats_data(
        partner_id=user.id,
        start_date=start_date,
        end_date=end_date,
    )

    os.makedirs("exports", exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    export_path = os.path.join("exports", f"my_settlement_{user.id}_{timestamp}.xlsx")

    export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    summary_rows = [
        {"指标": "报表名称", "数值": "我的结算报表"},
        {"指标": "导出时间", "数值": export_time},
        {"指标": "上传方", "数值": user.username},
        {"指标": "开始日期", "数值": start_date if start_date else "全部"},
        {"指标": "结束日期", "数值": end_date if end_date else "全部"},
        {"指标": "", "数值": ""},
        {"指标": "业务数据总条数", "数值": stats["total_records"]},
        {"指标": "总积分金额", "数值": stats["total_points"]},
        {"指标": "已通过结算条数", "数值": stats["approved_settlement_count"]},
        {"指标": "已通过结算金额", "数值": stats["approved_settlement_points"]},
        {"指标": "结算服务费合计", "数值": stats["approved_settlement_receivable_fee"]},
    ]

    internal_columns = [
        "业务单号",
        "姓名",
        "手机号",
        "车牌号",
        "积分金额",
        "银行卡号",
        "下游服务费率",
        "应收服务费",
        "导入时间",
    ]

    customer_column_names = {
        "下游服务费率": "结算费率",
        "应收服务费": "结算服务费",
    }

    detail_df = pd.DataFrame(stats["rows"])

    if detail_df.empty:
        detail_df = pd.DataFrame(
            columns=[
                "业务单号",
                "姓名",
                "手机号",
                "车牌号",
                "积分金额",
                "银行卡号",
                "结算费率",
                "结算服务费",
                "导入时间",
            ]
        )
    else:
        detail_df = detail_df[internal_columns]
        detail_df = detail_df.rename(columns=customer_column_names)

    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="汇总", index=False)
        detail_df.to_excel(writer, sheet_name="明细", index=False)

        format_excel_file(writer)

    return FileResponse(
        export_path,
        filename=f"我的结算报表_{timestamp}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/logout")
def logout():
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("user_id")
    return response


@app.get("/health")
def health():
    return {"status": "ok"}